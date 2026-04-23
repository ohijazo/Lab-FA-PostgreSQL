import io
import json
import urllib.request
import urllib.error
from datetime import date, datetime
from html import escape
from functools import wraps
from flask import Blueprint, jsonify, request, abort, session, send_file, current_app
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from app import db
from app.models import Analisi, EditLock, EmailLog, TipusAnalisi, User
from app.i18n import t

bp = Blueprint("analisis", __name__)


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "email" not in session:
            return jsonify({"error": t('no_autenticat')}), 401
        return f(*args, **kwargs)
    return decorated


def write_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "email" not in session:
            return jsonify({"error": t('no_autenticat')}), 401
        if session.get("role") == "viewer":
            return jsonify({"error": t('acces_lectura')}), 403
        return f(*args, **kwargs)
    return decorated


def _get_tipus_or_404(slug):
    tp = TipusAnalisi.query.filter_by(slug=slug).first()
    if not tp:
        abort(404, description=t('tipus_no_trobat', slug=slug))
    return tp


# --------------- Cerca per codi (escàner) ---------------

@bp.route("/api/analisis/find-by-codi", methods=["GET"])
@login_required
def find_by_codi():
    codi = request.args.get("codi", "").strip()
    if not codi:
        return jsonify({"error": t('codi_obligatori')}), 400
    analisi = Analisi.query.filter(
        Analisi.dades["codi"].as_string() == codi
    ).first()
    if not analisi:
        return jsonify({"error": t('no_trobat')}), 404
    return jsonify({"id": analisi.id, "tipus": analisi.tipus})


# --------------- Config endpoints ---------------

@bp.route("/api/tipus", methods=["GET"])
@login_required
def llistar_tipus():
    tots = TipusAnalisi.query.order_by(TipusAnalisi.nom).all()
    return jsonify([t.to_dict() for t in tots])


@bp.route("/api/tipus/<slug>/config", methods=["GET"])
@login_required
def config_tipus(slug):
    t = _get_tipus_or_404(slug)
    return jsonify(t.to_config())


# --------------- CRUD endpoints ---------------

@bp.route("/api/analisis/<slug>", methods=["GET"])
@login_required
def llistar(slug):
    t = _get_tipus_or_404(slug)

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 25, type=int)
    q = request.args.get("q", "").strip().lower()
    sort = request.args.get("sort", "").strip()
    sort_dir = request.args.get("sort_dir", "desc").strip().lower()

    query = Analisi.query.filter_by(tipus=slug)

    if q:
        query = query.filter(
            Analisi.dades.cast(db.Text).ilike(f"%{q}%")
        )

    # Advanced filters: f_{name}=val, f_{name}_from=val, f_{name}_to=val
    for key, val in request.args.items():
        if not key.startswith("f_") or not val.strip():
            continue
        val = val.strip()
        if key.endswith("_from"):
            field = key[2:-5]  # strip f_ and _from
            query = query.filter(Analisi.dades[field].as_string() >= val)
        elif key.endswith("_to"):
            field = key[2:-3]  # strip f_ and _to
            query = query.filter(Analisi.dades[field].as_string() <= val)
        else:
            field = key[2:]  # strip f_
            query = query.filter(Analisi.dades[field].as_string() == val)

    total = query.count()

    if sort:
        # Determine field type to sort numerically when appropriate
        camp_type = None
        for seccio in t.seccions:
            for camp in seccio.camps:
                if camp.name == sort:
                    camp_type = camp.type
                    break
            if camp_type:
                break

        if camp_type == "number":
            sort_expr = Analisi.dades[sort].as_float()
        else:
            sort_expr = Analisi.dades[sort].as_string()

        if sort_dir == "desc":
            query = query.order_by(sort_expr.desc().nullslast())
        else:
            query = query.order_by(sort_expr.asc().nullsfirst())
    else:
        query = query.order_by(Analisi.id.desc())

    analisis = query.offset((page - 1) * per_page).limit(per_page).all()

    return jsonify({
        "items": [a.to_dict() for a in analisis],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page if per_page else 1,
    })


def _build_export_sheet(ws, tipus_obj, analisis_list):
    """Build a worksheet with merged section headers (row 1) and camp labels (row 2)."""
    seccions_ordered = sorted(tipus_obj.seccions, key=lambda s: (s.ordre or 0))

    # Build structure: list of (seccio_titol, [(camp_name, camp_label), ...])
    seccio_camps = []
    for seccio in seccions_ordered:
        camps = sorted(seccio.camps, key=lambda c: (c.ordre or 0))
        if camps:
            seccio_camps.append((seccio.titol, [(c.name, c.label) for c in camps]))

    # Fixed columns before sections
    fixed_cols = 2  # ID, Data creació

    # Row 1: section headers (merged)
    # Row 2: camp labels
    section_header = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    # Write fixed column headers spanning both rows
    ws.cell(row=1, column=1, value="ID").font = section_header
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(row=1, column=1).alignment = center

    ws.cell(row=1, column=2, value="Data creació").font = section_header
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.cell(row=1, column=2).alignment = center

    col = fixed_cols + 1  # next column (1-indexed)
    all_camp_names = []
    for titol, camps in seccio_camps:
        num_camps = len(camps)
        # Merge section title across its camps in row 1
        ws.cell(row=1, column=col, value=titol).font = section_header
        ws.cell(row=1, column=col).alignment = center
        if num_camps > 1:
            ws.merge_cells(
                start_row=1, start_column=col,
                end_row=1, end_column=col + num_camps - 1,
            )
        # Write camp labels in row 2
        for i, (name, label) in enumerate(camps):
            ws.cell(row=2, column=col + i, value=label)
            all_camp_names.append(name)
        col += num_camps

    # Data rows starting from row 3
    for a in analisis_list:
        dades = a.dades if isinstance(a.dades, dict) else {}
        row = [
            a.id,
            a.created_at.strftime("%Y-%m-%d %H:%M") if a.created_at else "",
        ]
        for name in all_camp_names:
            row.append(dades.get(name, ""))
        ws.append(row)


@bp.route("/api/analisis/<slug>/export", methods=["GET"])
@login_required
def exportar(slug):
    t = _get_tipus_or_404(slug)
    q = request.args.get("q", "").strip().lower()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    all_fields = request.args.get("all_fields", "0").strip() == "1"

    query = Analisi.query.filter_by(tipus=slug)
    if q:
        query = query.filter(Analisi.dades.cast(db.Text).ilike(f"%{q}%"))
    if date_from:
        query = query.filter(Analisi.dades["data"].as_string() >= date_from)
    if date_to:
        query = query.filter(Analisi.dades["data"].as_string() <= date_to)
    query = query.order_by(Analisi.id.desc())
    analisis = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = t.nom

    if all_fields:
        _build_export_sheet(ws, t, analisis)
    else:
        # Simple export with columnes_llista only (no section headers)
        columnes = t.columnes_llista or []
        seccions_ordered = sorted(t.seccions, key=lambda s: (s.ordre or 0))
        camp_labels = {}
        for seccio in seccions_ordered:
            for camp in seccio.camps:
                camp_labels[camp.name] = camp.label
        headers = ["ID", "Data creació"] + [camp_labels.get(c, c) for c in columnes]
        ws.append(headers)
        for a in analisis:
            dades = a.dades if isinstance(a.dades, dict) else {}
            row = [a.id, a.created_at.strftime("%Y-%m-%d %H:%M") if a.created_at else ""]
            for c in columnes:
                row.append(dades.get(c, ""))
            ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{slug}_{date.today().isoformat()}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@bp.route("/api/analisis/export-multi", methods=["GET"])
@login_required
def exportar_multi():
    slugs = request.args.getlist("tipus")
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    if not slugs:
        abort(400, description=t('cal_seleccionar_tipus'))

    tipus_list = TipusAnalisi.query.filter(TipusAnalisi.slug.in_(slugs)).all()
    if not tipus_list:
        abort(404, description=t('cap_tipus_trobat'))

    wb = Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    for t in sorted(tipus_list, key=lambda x: x.nom):
        query = Analisi.query.filter_by(tipus=t.slug)
        if date_from:
            query = query.filter(Analisi.dades["data"].as_string() >= date_from)
        if date_to:
            query = query.filter(Analisi.dades["data"].as_string() <= date_to)
        query = query.order_by(Analisi.id.desc())
        analisis = query.all()

        # Sheet title max 31 chars (Excel limit)
        ws = wb.create_sheet(title=t.nom[:31])
        _build_export_sheet(ws, t, analisis)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"analisis_{date.today().isoformat()}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def _resolve_user_name(value):
    """If value looks like an email, try to find the user's nom."""
    if not value or "@" not in value:
        return value
    u = User.query.filter_by(email=value).first()
    return u.nom if u and u.nom else value


@bp.route("/api/analisis/<slug>/<int:id>", methods=["GET"])
@login_required
def detall(slug, id):
    a = db.get_or_404(Analisi, id)
    if a.tipus != slug:
        abort(404)
    result = a.to_dict()
    result["created_by"] = _resolve_user_name(result.get("created_by"))
    result["updated_by"] = _resolve_user_name(result.get("updated_by"))
    return jsonify(result)


@bp.route("/api/analisis/<slug>", methods=["POST"])
@write_required
def crear(slug):
    _get_tipus_or_404(slug)

    data = request.get_json()
    for key in ("id", "created_at", "updated_at", "tipus", "created_by", "updated_by"):
        data.pop(key, None)

    # Validate unique codi within this tipus
    codi = data.get("codi", "").strip()
    if codi:
        existing = Analisi.query.filter_by(tipus=slug).filter(
            Analisi.dades["codi"].as_string() == codi
        ).first()
        if existing:
            return jsonify({"error": t('ja_existeix_codi', codi=codi)}), 409

    a = Analisi(tipus=slug, created_by=session.get("nom"), updated_by=session.get("nom"))
    a.set_dades(data)
    db.session.add(a)
    db.session.commit()
    return jsonify(a.to_dict()), 201


@bp.route("/api/analisis/<slug>/<int:id>", methods=["PUT"])
@write_required
def editar(slug, id):
    a = db.get_or_404(Analisi, id)
    if a.tipus != slug:
        abort(404)

    data = request.get_json()

    # Optimistic locking: check _expected_updated_at
    expected = data.pop("_expected_updated_at", None)
    if expected and a.updated_at:
        actual = a.updated_at.isoformat()
        if expected != actual:
            return jsonify({
                "error": "conflict",
                "message": t('conflicte_concurrencia', user=a.updated_by or 'un altre usuari'),
                "updated_by": a.updated_by,
                "updated_at": actual,
                "current": a.to_dict(),
            }), 409

    for key in ("id", "created_at", "updated_at", "tipus", "created_by", "updated_by"):
        data.pop(key, None)

    # Validate unique codi within this tipus (excluding current record)
    codi = data.get("codi", "").strip()
    if codi:
        existing = Analisi.query.filter_by(tipus=slug).filter(
            Analisi.dades["codi"].as_string() == codi,
            Analisi.id != id,
        ).first()
        if existing:
            return jsonify({"error": t('ja_existeix_codi', codi=codi)}), 409

    a.set_dades(data)
    a.updated_at = datetime.utcnow()
    a.updated_by = session.get("nom")
    db.session.commit()
    return jsonify(a.to_dict())


@bp.route("/api/analisis/<slug>/<int:id>", methods=["DELETE"])
@write_required
def eliminar(slug, id):
    a = db.get_or_404(Analisi, id)
    if a.tipus != slug:
        abort(404)
    # Clean up any locks for this analisi
    EditLock.query.filter_by(analisi_id=id).delete()
    db.session.delete(a)
    db.session.commit()
    return jsonify({"ok": True})


# --------------- Edit lock (presence) ---------------

@bp.route("/api/analisis/<slug>/<int:id>/lock", methods=["POST"])
@write_required
def acquire_lock(slug, id):
    a = db.get_or_404(Analisi, id)
    if a.tipus != slug:
        abort(404)

    email = session.get("email")

    # Clean expired locks
    all_locks = EditLock.query.filter_by(analisi_id=id).all()
    for lock in all_locks:
        if lock.is_expired():
            db.session.delete(lock)

    # Check for existing lock by another user
    other_lock = EditLock.query.filter(
        EditLock.analisi_id == id,
        EditLock.user_email != email,
    ).first()

    # Upsert own lock
    own_lock = EditLock.query.filter_by(analisi_id=id, user_email=email).first()
    if own_lock:
        own_lock.locked_at = datetime.utcnow()
    else:
        own_lock = EditLock(analisi_id=id, user_email=email)
        db.session.add(own_lock)

    db.session.commit()

    result = {
        "ok": True,
        "lock": own_lock.to_dict(),
        "updated_at": a.updated_at.isoformat() if a.updated_at else None,
        "updated_by": a.updated_by,
    }
    if other_lock:
        result["other_user"] = other_lock.to_dict()
    return jsonify(result)


@bp.route("/api/analisis/<slug>/<int:id>/lock", methods=["DELETE"])
@write_required
def release_lock(slug, id):
    email = session.get("email")
    EditLock.query.filter_by(analisi_id=id, user_email=email).delete()
    db.session.commit()
    return jsonify({"ok": True})


# --------------- Send email (via Power Automate webhook) ---------------

@bp.route("/api/analisis/<slug>/<int:id>/email", methods=["POST"])
@login_required
def enviar_email(slug, id):
    webhook_url = current_app.config.get("POWER_AUTOMATE_WEBHOOK_URL", "")
    if not webhook_url:
        return jsonify({"error": t('email_no_configurat')}), 400

    user = User.query.filter_by(email=session.get("email")).first()
    if not user:
        return jsonify({"error": t('no_autenticat')}), 401

    data = request.get_json()
    destinatari = (data.get("destinatari") or "").strip()
    if not destinatari:
        return jsonify({"error": t('email_destinatari_obligatori')}), 400

    assumpte = (data.get("assumpte") or "").strip()

    tp = _get_tipus_or_404(slug)
    a = db.get_or_404(Analisi, id)
    if a.tipus != slug:
        abort(404)

    dades = a.dades if isinstance(a.dades, dict) else {}
    seccions = sorted(tp.seccions, key=lambda s: (s.ordre or 0))

    # Determine first column value for reference
    cols = tp.columnes_llista or []
    ref_value = dades.get(cols[0], f"#{a.id}") if cols else f"#{a.id}"

    def _format_date_value(val):
        """Convert YYYY-MM-DD to DD-MM-YYYY if it looks like a date."""
        if isinstance(val, str) and len(val) >= 10 and val[4:5] == "-" and val[7:8] == "-":
            try:
                return datetime.strptime(val[:10], "%Y-%m-%d").strftime("%d-%m-%Y")
            except ValueError:
                pass
        return val

    ref_value = _format_date_value(ref_value)

    # Build HTML body
    data_str = datetime.utcnow().strftime("%d/%m/%Y %H:%M")
    html_parts = []
    html_parts.append('<div style="background-color:#f4f6f8;padding:32px 0;font-family:Arial,Helvetica,sans-serif">')
    html_parts.append('<div style="max-width:680px;margin:0 auto">')

    # 1) Introductory text OUTSIDE the card, before everything
    html_parts.append(f'<p style="color:#333;font-size:15px;line-height:1.6;margin:0 0 20px 0;padding:0 4px">{t("email_intro", nom=escape(user.nom), tipus=escape(tp.nom), ref=escape(str(ref_value)))}</p>')

    # 2) Card with header + data
    html_parts.append('<div style="background:#ffffff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.08)">')

    # Header bar
    html_parts.append('<div style="background:#1a56db;padding:20px 32px">')
    html_parts.append('<span style="color:#ffffff;font-size:20px;font-weight:600">Lab FA</span>')
    html_parts.append('</div>')

    # Body content
    html_parts.append('<div style="padding:24px 32px">')

    for seccio in seccions:
        camps = sorted(seccio.camps, key=lambda c: (c.ordre or 0))
        if not camps:
            continue
        html_parts.append(f'<h3 style="color:#374151;font-size:13px;font-weight:600;margin:20px 0 8px 0;text-transform:uppercase;letter-spacing:0.5px">{escape(seccio.titol)}</h3>')
        html_parts.append('<table style="width:100%;border-collapse:collapse;margin-bottom:16px">')
        for i, camp in enumerate(camps):
            val = dades.get(camp.name, "")
            if val is None:
                val = ""
            if camp.type == "date":
                val = _format_date_value(val)
            elif camp.type == "checkbox":
                val = t("email_si") if val else t("email_no")
            bg = "#f9fafb" if i % 2 == 0 else "#ffffff"
            html_parts.append(
                f'<tr style="background:{bg}">'
                f'<td style="padding:8px 12px;border-bottom:1px solid #e5e7eb;color:#6b7280;font-size:13px;width:40%">{escape(camp.label)}</td>'
                f'<td style="padding:8px 12px;border-bottom:1px solid #e5e7eb;color:#111827;font-size:13px;font-weight:500">{escape(str(val))}</td>'
                f'</tr>'
            )
        html_parts.append('</table>')

    # Footer inside card
    html_parts.append('<div style="margin-top:24px;padding-top:14px;border-top:1px solid #e5e7eb">')
    html_parts.append(f'<p style="font-size:12px;color:#9ca3af;margin:0 0 4px 0">{t("email_peu", nom=user.nom, data=data_str)}</p>')
    html_parts.append(f'<p style="font-size:11px;color:#d1d5db;margin:0">{t("email_automatic")}</p>')
    html_parts.append('</div>')

    html_parts.append('</div>')  # end padding div
    html_parts.append('</div>')  # end card
    html_parts.append('</div>')  # end centering div
    html_parts.append('</div>')  # end outer wrapper

    html_body = "\n".join(html_parts)

    if not assumpte:
        assumpte = tp.nom

    payload = json.dumps({
        "destinatari": destinatari,
        "assumpte": assumpte,
        "cos_html": html_body,
        "remitent_nom": user.nom,
        "remitent_email": user.email,
    }).encode("utf-8")

    try:
        req = urllib.request.Request(
            webhook_url,
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            if resp.status >= 400:
                return jsonify({"error": t('email_error', error=f"HTTP {resp.status}")}), 500
    except urllib.error.HTTPError as e:
        return jsonify({"error": t('email_error', error=f"HTTP {e.code}")}), 500
    except Exception as e:
        return jsonify({"error": t('email_error', error=str(e))}), 500

    # Save email log
    log = EmailLog(
        analisi_id=a.id,
        tipus_slug=slug,
        destinatari=destinatari,
        assumpte=assumpte,
        enviat_per=user.email,
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({"ok": True, "message": t('email_enviat')})


@bp.route("/api/analisis/<slug>/<int:id>/emails", methods=["GET"])
@login_required
def llistar_emails(slug, id):
    a = db.get_or_404(Analisi, id)
    if a.tipus != slug:
        abort(404)
    logs = EmailLog.query.filter_by(analisi_id=id).order_by(EmailLog.enviat_at.desc()).all()
    return jsonify([l.to_dict() for l in logs])
