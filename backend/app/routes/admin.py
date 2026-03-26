import re
import json
from datetime import datetime
from functools import wraps
from flask import Blueprint, jsonify, request, abort, session
from io import BytesIO
from flask import send_file
from openpyxl import Workbook, load_workbook
from app import db
from app.models import TipusAnalisi, Seccio, Camp, User, Analisi
from app.i18n import t as tr

bp = Blueprint("admin", __name__)


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "email" not in session:
            return jsonify({"error": tr('no_autenticat')}), 401
        if session.get("role") != "admin":
            return jsonify({"error": tr('acces_denegat')}), 403
        return f(*args, **kwargs)
    return decorated


def _slugify(text):
    text = text.lower().strip()
    text = re.sub(r"[àáâã]", "a", text)
    text = re.sub(r"[èéêë]", "e", text)
    text = re.sub(r"[ìíîï]", "i", text)
    text = re.sub(r"[òóôõ]", "o", text)
    text = re.sub(r"[ùúûü]", "u", text)
    text = re.sub(r"[ç]", "c", text)
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


# ===================== TIPUS ANALISI =====================

@bp.route("/api/admin/tipus", methods=["GET"])
@admin_required
def llistar_tipus():
    tots = TipusAnalisi.query.order_by(TipusAnalisi.nom).all()
    return jsonify([t.to_dict() for t in tots])


@bp.route("/api/admin/tipus", methods=["POST"])
@admin_required
def crear_tipus():
    data = request.get_json()
    nom = data.get("nom", "").strip()
    if not nom:
        abort(400, description=tr('nom_obligatori'))

    slug = data.get("slug") or _slugify(nom)
    if TipusAnalisi.query.filter_by(slug=slug).first():
        abort(409, description=tr('ja_existeix_slug', slug=slug))

    t = TipusAnalisi(
        nom=nom,
        slug=slug,
        descripcio=data.get("descripcio", ""),
    )
    if "columnes_llista" in data:
        t.set_columnes_llista(data["columnes_llista"])
    db.session.add(t)
    db.session.commit()
    return jsonify(t.to_dict()), 201


@bp.route("/api/admin/tipus/<int:id>", methods=["GET"])
@admin_required
def detall_tipus(id):
    t = db.get_or_404(TipusAnalisi, id)
    return jsonify(t.to_config())


@bp.route("/api/admin/tipus/<int:id>", methods=["PUT"])
@admin_required
def editar_tipus(id):
    t = db.get_or_404(TipusAnalisi, id)
    data = request.get_json()
    if "nom" in data:
        t.nom = data["nom"].strip()
    if "slug" in data:
        new_slug = data["slug"].strip()
        if new_slug != t.slug:
            # Prohibir canvi de slug si hi ha analisis vinculades
            n_analisis = Analisi.query.filter_by(tipus=t.slug).count()
            if n_analisis > 0:
                return jsonify({"error": tr('no_canviar_slug', n=n_analisis)}), 400
            existing = TipusAnalisi.query.filter_by(slug=new_slug).first()
            if existing:
                return jsonify({"error": tr('slug_ja_existeix', slug=new_slug)}), 409
            t.slug = new_slug
    if "descripcio" in data:
        t.descripcio = data["descripcio"]
    if "columnes_llista" in data:
        t.set_columnes_llista(data["columnes_llista"])
    db.session.commit()
    return jsonify(t.to_dict())


@bp.route("/api/admin/tipus/<int:id>", methods=["DELETE"])
@admin_required
def eliminar_tipus(id):
    t = db.get_or_404(TipusAnalisi, id)
    n = Analisi.query.filter_by(tipus=t.slug).count()
    if n > 0:
        return jsonify({"error": tr('no_eliminar_analisis', n=n)}), 400
    db.session.delete(t)
    db.session.commit()
    return jsonify({"ok": True})


# ===================== SECCIONS =====================

@bp.route("/api/admin/tipus/<int:tipus_id>/seccions", methods=["GET"])
@admin_required
def llistar_seccions(tipus_id):
    db.get_or_404(TipusAnalisi, tipus_id)
    seccions = Seccio.query.filter_by(tipus_id=tipus_id).order_by(Seccio.ordre).all()
    return jsonify([s.to_dict() for s in seccions])


@bp.route("/api/admin/tipus/<int:tipus_id>/seccions", methods=["POST"])
@admin_required
def crear_seccio(tipus_id):
    db.get_or_404(TipusAnalisi, tipus_id)
    data = request.get_json()
    titol = data.get("titol", "").strip()
    if not titol:
        abort(400, description=tr('titol_obligatori'))

    max_ordre = db.session.query(db.func.max(Seccio.ordre)).filter_by(tipus_id=tipus_id).scalar() or 0

    s = Seccio(
        tipus_id=tipus_id,
        titol=titol,
        ordre=data.get("ordre", max_ordre + 1),
    )
    db.session.add(s)
    db.session.commit()
    return jsonify(s.to_dict()), 201


@bp.route("/api/admin/tipus/<int:tipus_id>/seccions/reorder", methods=["PUT"])
@admin_required
def reordenar_seccions(tipus_id):
    db.get_or_404(TipusAnalisi, tipus_id)
    data = request.get_json()
    ordered_ids = data.get("order", [])
    for idx, sid in enumerate(ordered_ids):
        s = db.session.get(Seccio, sid)
        if s and s.tipus_id == tipus_id:
            s.ordre = idx
    db.session.commit()
    seccions = Seccio.query.filter_by(tipus_id=tipus_id).order_by(Seccio.ordre).all()
    return jsonify([s.to_dict() for s in seccions])


@bp.route("/api/admin/seccions/<int:id>", methods=["PUT"])
@admin_required
def editar_seccio(id):
    s = db.get_or_404(Seccio, id)
    data = request.get_json()
    if "titol" in data:
        s.titol = data["titol"].strip()
    if "ordre" in data:
        s.ordre = data["ordre"]
    db.session.commit()
    return jsonify(s.to_dict())


@bp.route("/api/admin/seccions/<int:id>", methods=["DELETE"])
@admin_required
def eliminar_seccio(id):
    s = db.get_or_404(Seccio, id)
    t = db.get_or_404(TipusAnalisi, s.tipus_id)
    n = Analisi.query.filter_by(tipus=t.slug).count()
    if n > 0:
        return jsonify({"error": tr('no_eliminar_seccio', n=n, nom=t.nom)}), 400
    db.session.delete(s)
    db.session.commit()
    return jsonify({"ok": True})


# ===================== CAMPS =====================

@bp.route("/api/admin/seccions/<int:seccio_id>/camps", methods=["GET"])
@admin_required
def llistar_camps(seccio_id):
    db.get_or_404(Seccio, seccio_id)
    camps = Camp.query.filter_by(seccio_id=seccio_id).order_by(Camp.ordre).all()
    return jsonify([c.to_dict() for c in camps])


@bp.route("/api/admin/seccions/<int:seccio_id>/camps", methods=["POST"])
@admin_required
def crear_camp(seccio_id):
    db.get_or_404(Seccio, seccio_id)
    data = request.get_json()

    name = data.get("name", "").strip()
    label = data.get("label", "").strip()
    if not name or not label:
        abort(400, description=tr('name_label_obligatoris'))

    max_ordre = db.session.query(db.func.max(Camp.ordre)).filter_by(seccio_id=seccio_id).scalar() or 0

    c = Camp(
        seccio_id=seccio_id,
        name=name,
        label=label,
        type=data.get("type", "text"),
        required=data.get("required", False),
        ordre=data.get("ordre", max_ordre + 1),
        grup=data.get("grup", ""),
        opcions=data.get("opcions", []),
        alerta_min=data.get("alerta_min"),
        alerta_max=data.get("alerta_max"),
        alerta_color_min=data.get("alerta_color_min"),
        alerta_color_max=data.get("alerta_color_max"),
    )
    db.session.add(c)
    db.session.commit()
    return jsonify(c.to_dict()), 201


@bp.route("/api/admin/seccions/<int:seccio_id>/camps/reorder", methods=["PUT"])
@admin_required
def reordenar_camps(seccio_id):
    db.get_or_404(Seccio, seccio_id)
    data = request.get_json()
    ordered_ids = data.get("order", [])
    for idx, cid in enumerate(ordered_ids):
        c = db.session.get(Camp, cid)
        if c and c.seccio_id == seccio_id:
            c.ordre = idx
    db.session.commit()
    camps = Camp.query.filter_by(seccio_id=seccio_id).order_by(Camp.ordre).all()
    return jsonify([c.to_dict() for c in camps])


@bp.route("/api/admin/camps/<int:id>", methods=["PUT"])
@admin_required
def editar_camp(id):
    c = db.get_or_404(Camp, id)
    data = request.get_json()
    if "name" in data:
        c.name = data["name"].strip()
    if "label" in data:
        c.label = data["label"].strip()
    if "type" in data:
        c.type = data["type"]
    if "required" in data:
        c.required = data["required"]
    if "ordre" in data:
        c.ordre = data["ordre"]
    if "grup" in data:
        c.grup = data["grup"]
    if "opcions" in data:
        c.opcions = data["opcions"]
    if "alerta_min" in data:
        c.alerta_min = data["alerta_min"]
    if "alerta_max" in data:
        c.alerta_max = data["alerta_max"]
    if "alerta_color_min" in data:
        c.alerta_color_min = data["alerta_color_min"]
    if "alerta_color_max" in data:
        c.alerta_color_max = data["alerta_color_max"]
    db.session.commit()
    return jsonify(c.to_dict())


@bp.route("/api/admin/camps/<int:id>", methods=["DELETE"])
@admin_required
def eliminar_camp(id):
    c = db.get_or_404(Camp, id)
    s = db.get_or_404(Seccio, c.seccio_id)
    t = db.get_or_404(TipusAnalisi, s.tipus_id)
    n = Analisi.query.filter_by(tipus=t.slug).count()
    if n > 0:
        return jsonify({"error": tr('no_eliminar_camp', n=n, nom=t.nom)}), 400
    db.session.delete(c)
    db.session.commit()
    return jsonify({"ok": True})


# ===================== PLANTILLA EXCEL =====================

@bp.route("/api/admin/tipus/<int:id>/plantilla", methods=["GET"])
@admin_required
def descarregar_plantilla(id):
    t = db.get_or_404(TipusAnalisi, id)

    all_labels = []
    for seccio in t.seccions:
        for camp in seccio.camps:
            all_labels.append(camp.label)

    if not all_labels:
        return jsonify({"error": tr('tipus_sense_camps')}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = t.nom
    ws.append(all_labels)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"plantilla_{t.slug}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ===================== DUPLICAR TIPUS =====================

@bp.route("/api/admin/tipus/<int:id>/duplicar", methods=["POST"])
@admin_required
def duplicar_tipus(id):
    original = db.get_or_404(TipusAnalisi, id)

    base_nom = f"{original.nom} (copia)"
    base_slug = _slugify(base_nom)

    # Assegurar slug unic
    slug = base_slug
    counter = 2
    while TipusAnalisi.query.filter_by(slug=slug).first():
        slug = f"{base_slug}_{counter}"
        counter += 1

    nou = TipusAnalisi(
        nom=base_nom,
        slug=slug,
        descripcio=original.descripcio,
        columnes_llista=original.get_columnes_llista(),
    )
    db.session.add(nou)
    db.session.flush()

    for seccio in original.seccions:
        nova_seccio = Seccio(
            tipus_id=nou.id,
            titol=seccio.titol,
            ordre=seccio.ordre,
        )
        db.session.add(nova_seccio)
        db.session.flush()

        for camp in seccio.camps:
            nou_camp = Camp(
                seccio_id=nova_seccio.id,
                name=camp.name,
                label=camp.label,
                type=camp.type,
                required=camp.required,
                ordre=camp.ordre,
                grup=camp.grup,
                opcions=camp.opcions,
                alerta_min=camp.alerta_min,
                alerta_max=camp.alerta_max,
                alerta_color_min=camp.alerta_color_min,
                alerta_color_max=camp.alerta_color_max,
            )
            db.session.add(nou_camp)

    db.session.commit()
    return jsonify(nou.to_dict()), 201


# ===================== ESTADISTIQUES =====================

@bp.route("/api/admin/estadistiques", methods=["GET"])
@admin_required
def estadistiques():
    tots_tipus = TipusAnalisi.query.order_by(TipusAnalisi.nom).all()
    result = []
    for t in tots_tipus:
        count = Analisi.query.filter_by(tipus=t.slug).count()
        ultima = db.session.query(db.func.max(Analisi.created_at)).filter_by(tipus=t.slug).scalar()
        result.append({
            "id": t.id,
            "nom": t.nom,
            "slug": t.slug,
            "total_analisis": count,
            "ultima_analisi": ultima.isoformat() if ultima else None,
        })
    return jsonify(result)


# ===================== IMPORTACIO EXCEL =====================

@bp.route("/api/admin/tipus/<int:id>/import", methods=["POST"])
@admin_required
def importar_excel(id):
    t = db.get_or_404(TipusAnalisi, id)

    if "file" not in request.files:
        return jsonify({"error": tr('no_fitxer')}), 400

    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": tr('fitxer_format')}), 400

    # Load camp config
    all_camps = []
    for seccio in t.seccions:
        for camp in seccio.camps:
            all_camps.append(camp)

    # Build lookup: lowercase label -> camp, lowercase name -> camp
    label_map = {}
    name_map = {}
    for c in all_camps:
        label_map[c.label.strip().lower()] = c
        name_map[c.name.strip().lower()] = c

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return jsonify({"error": tr('no_llegir_excel')}), 400

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({"error": tr('fitxer_sense_dades')}), 400

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    # Match columns to camps
    col_mapping = {}  # col_index -> Camp object
    unrecognized = []
    for idx, header in enumerate(headers):
        h_lower = header.lower()
        if h_lower in label_map:
            col_mapping[idx] = label_map[h_lower]
        elif h_lower in name_map:
            col_mapping[idx] = name_map[h_lower]
        elif header:
            unrecognized.append(header)

    if not col_mapping:
        return jsonify({"error": tr('cap_columna_coincideix'), "columnes_no_reconegudes": unrecognized}), 400

    # Find required camps
    required_camps = {c.name for c in all_camps if c.required}

    # Get existing codis for dedup
    existing_codis = set()
    existing_analisis = Analisi.query.filter_by(tipus=t.slug).all()
    for a in existing_analisis:
        d = a.get_dades()
        if "codi" in d and d["codi"]:
            existing_codis.add(str(d["codi"]).strip().lower())

    importats = 0
    saltats = 0
    errors = []
    user_nom = session.get("nom", session.get("email", "import"))

    for row_num, row in enumerate(rows[1:], start=2):
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        dades = {}
        row_errors = []

        for col_idx, camp in col_mapping.items():
            val = row[col_idx] if col_idx < len(row) else None

            if val is None or str(val).strip() == "":
                if camp.name in required_camps:
                    row_errors.append(tr('camp_obligatori_buit', label=camp.label))
                continue

            # Type conversion
            if camp.type == "number":
                try:
                    val = float(val)
                    if val == int(val):
                        val = int(val)
                except (ValueError, TypeError):
                    row_errors.append(tr('valor_no_numeric', label=camp.label, val=val))
                    continue
            elif camp.type == "checkbox":
                if isinstance(val, bool):
                    val = val
                elif isinstance(val, str):
                    val = val.strip().lower() in ("true", "si", "sí", "1", "yes", "x")
                else:
                    val = bool(val)
            elif camp.type == "date":
                if isinstance(val, datetime):
                    val = val.strftime("%Y-%m-%d")
                else:
                    val = str(val).strip()
            else:
                val = str(val).strip()

            dades[camp.name] = val

        if row_errors:
            errors.append({"fila": row_num, "motiu": "; ".join(row_errors)})
            continue

        # Check duplicate codi
        if "codi" in dades and dades["codi"]:
            codi_lower = str(dades["codi"]).strip().lower()
            if codi_lower in existing_codis:
                saltats += 1
                continue
            existing_codis.add(codi_lower)

        analisi = Analisi(
            tipus=t.slug,
            dades=dades,
            created_by=user_nom,
            updated_by=user_nom,
        )
        db.session.add(analisi)
        importats += 1

    db.session.commit()
    wb.close()

    return jsonify({
        "importats": importats,
        "saltats": saltats,
        "errors": errors,
        "columnes_reconegudes": [col_mapping[i].label for i in sorted(col_mapping.keys())],
        "columnes_no_reconegudes": unrecognized,
    })


# ===================== USUARIS =====================

@bp.route("/api/admin/users", methods=["GET"])
@admin_required
def llistar_users():
    users = User.query.order_by(User.email).all()
    return jsonify([u.to_dict() for u in users])


@bp.route("/api/admin/users", methods=["POST"])
@admin_required
def crear_user():
    data = request.get_json()
    email = data.get("email", "").strip().lower()
    nom = data.get("nom", "").strip()
    password = data.get("password", "").strip()
    role = data.get("role", "user")

    if not email or not password:
        abort(400, description=tr('email_password_obligatoris'))
    if not nom:
        abort(400, description=tr('nom_obligatori_user'))
    if role not in ("admin", "user", "viewer"):
        abort(400, description=tr('role_invalid'))
    if User.query.filter_by(email=email).first():
        abort(409, description=tr('ja_existeix_email', email=email))

    u = User(email=email, nom=nom, role=role)
    u.set_password(password)
    db.session.add(u)
    db.session.commit()
    return jsonify(u.to_dict()), 201


@bp.route("/api/admin/users/<int:id>", methods=["PUT"])
@admin_required
def editar_user(id):
    u = db.get_or_404(User, id)
    data = request.get_json()

    if "email" in data:
        new_email = data["email"].strip().lower()
        existing = User.query.filter_by(email=new_email).first()
        if existing and existing.id != u.id:
            abort(409, description=tr('ja_existeix_email', email=new_email))
        u.email = new_email
    if "nom" in data:
        u.nom = data["nom"].strip()
    if "role" in data:
        if data["role"] not in ("admin", "user"):
            abort(400, description=tr('role_invalid'))
        u.role = data["role"]
    if "password" in data and data["password"].strip():
        u.set_password(data["password"].strip())
    if "email_from_name" in data:
        u.email_from_name = data["email_from_name"].strip()
    if "email_from_address" in data:
        u.email_from_address = data["email_from_address"].strip()
    if "email_smtp_password" in data:
        val = data["email_smtp_password"]
        if val and val != "••••••••":
            u.email_smtp_password = val.strip()

    db.session.commit()
    return jsonify(u.to_dict())


@bp.route("/api/admin/users/<int:id>", methods=["DELETE"])
@admin_required
def eliminar_user(id):
    u = db.get_or_404(User, id)
    if u.id == session.get("user_id"):
        abort(400, description=tr('no_eliminar_propi'))
    db.session.delete(u)
    db.session.commit()
    return jsonify({"ok": True})
